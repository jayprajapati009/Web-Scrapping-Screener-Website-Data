from ast import Continue
from time import process_time_ns
from typing import final
from bs4 import BeautifulSoup as bs
from numpy import append
import requests
import xlsxwriter
import os
import openpyxl

# Setting up the directory to save the excel file in the same folder.
os.chdir(r'C:\Users\jaypr\Desktop\Tech Stack\VSCodes\Web Scrapping\StockScrapping\Scrapping Screener Website Data\Scraped Data')


def find_data(link, row):
    print(link)
    # To fetch the html data from the website
    html_text = requests.get(link).text

    # Parsing the data using lxml Parser and Beautiful Soup Library
    soup = bs(html_text, 'lxml')

    ##### Company Name #####
    Company = soup.find(
        'h1', class_='margin-0').text

    ##### Balance Sheet Months #####
    borrowings_dates_list = soup.find(
        'section', id='balance-sheet', class_='card card-large').table.thead.tr.text.strip().split('\n')

    ##### Borrowings or Debts - Balance Sheet #####
    borrowings_values = soup.find(
        'section', id='balance-sheet', class_='card card-large').text
    val_i = borrowings_values.find('Borrowings')
    othli_i = borrowings_values.find('Other Liabilities')
    borrowings_values_list = borrowings_values[val_i +
                                               10: othli_i].strip().split('\n')

    ##### Share Capital - Balance Sheet #####
    shareCap = soup.find(
        'section', id='balance-sheet', class_='card card-large').text
    shcap_i = shareCap.find('Share Capital')
    Res = shareCap.find('Reserves')
    shareCap_list = shareCap[shcap_i+15: Res].strip().split('\n')

    ##### Reserves - Balance Sheet #####
    Reserves = soup.find(
        'section', id='balance-sheet', class_='card card-large').text
    shcap_i = Reserves.find('Reserves')
    Res = Reserves.find('Borrowings')
    Reserves_list = Reserves[shcap_i+15: Res].strip().split('\n')

    ##### Profit and Loss Months #####
    pldates_list = soup.find(
        'section', id='balance-sheet', class_='card card-large').table.thead.tr.text.strip().split('\n')

    ##### Profit Before Tax (PBT) - Profit and Loss #####
    pbt = soup.find(
        'section', id='profit-loss', class_='card card-large').text
    pbt_i = pbt.find('Profit before tax')
    tax_i = pbt.find('Tax %')
    pbt_list = pbt[pbt_i+25: tax_i].strip().split('\n')

    ##### Net Profit (PAT) - Profit and Loss #####
    pat = soup.find(
        'section', id='profit-loss', class_='card card-large').text
    pat_i = pat.find('Net Profit')
    tax_i = pat.find('EPS')
    pat_list = pat[pat_i+15: tax_i].strip().split('\n')

    ##### Sales - Profit and Loss #####
    sales = soup.find(
        'section', id='profit-loss', class_='card card-large').text
    sales_i = sales.find('Sales\xa0+')
    tax_i = sales.find('Expenses\xa0+')
    sales_list = sales[sales_i+10: tax_i].strip().split('\n')

    ##### Other Income - Profit and Loss #####
    othInc = soup.find(
        'section', id='profit-loss', class_='card card-large').text
    othInc_i = othInc.find('Other')
    tax_i = othInc.find('Interest')
    othInc_list = othInc[othInc_i+15: tax_i].strip().split('\n')

    # Create a workbook and add a worksheet
    workbook = xlsxwriter.Workbook('Data.xlsx')
    worksheet = workbook.add_worksheet()

    # Borrowings Coloum (B) width
    worksheet.set_column('A:A', 30)
    worksheet.set_column('B:B', 15)
    worksheet.set_column('D:D', 15)
    worksheet.set_column('G:G', 10)

    # Defined Formats
    bold = workbook.add_format({'bold': True})
    cen = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
    title = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'bold': True})
    merge_format = workbook.add_format({
        'bold': 1,
        'align': 'center',
        'valign': 'vcenter'})

    for slicingIndexb in range(len(borrowings_dates_list)):
        if "2017" in borrowings_dates_list[slicingIndexb]:
            break
        elif "2018" in borrowings_dates_list[slicingIndexb]:
            break
        elif "2019" in borrowings_dates_list[slicingIndexb]:
            break
        elif "2020" in borrowings_dates_list[slicingIndexb]:
            break
        elif "2021" in borrowings_dates_list[slicingIndexb]:
            break
        elif "2022" in borrowings_dates_list[slicingIndexb]:
            break
        else:
            slicingIndexb = slicingIndexb + 1

    for slicingIndex in range(len(pldates_list)):
        if "2017" in pldates_list[slicingIndex]:
            break
        elif "2018" in pldates_list[slicingIndex]:
            break
        elif "2019" in pldates_list[slicingIndex]:
            break
        elif "2020" in pldates_list[slicingIndex]:
            break
        elif "2021" in pldates_list[slicingIndex]:
            break
        elif "2022" in pldates_list[slicingIndex]:
            break
        else:
            slicingIndex = slicingIndex + 1

    updated_borrowings_dates_list = borrowings_dates_list[slicingIndexb:len(
        borrowings_dates_list)]
    updated_borrowings_values_list = borrowings_values_list[slicingIndexb:len(
        borrowings_values)]
    updated_shareCap_list = shareCap_list[slicingIndexb:len(shareCap_list)]
    updated_Reserves_list = Reserves_list[slicingIndexb:len(Reserves_list)]
    shareHoldersFund = [float(updated_shareCap_list[i]) + float(updated_Reserves_list[i])
                        for i in range(len(updated_shareCap_list))]

    updated_pldates_list = pldates_list[slicingIndex:len(pldates_list)]
    updated_pbt_list = pbt_list[slicingIndex:len(pbt_list)]
    updated_pat_list = pat_list[slicingIndex:len(pat_list)]
    updated_sales_listt = sales_list[slicingIndex:len(sales_list)]
    updated_othInc_list = othInc_list[slicingIndex:len(othInc_list)]
    totalRevenue = [float(updated_sales_listt[i]) + float(updated_othInc_list[i])
                    for i in range(len(updated_sales_listt))]

    # Headings Declarations
    worksheet.merge_range('B1:G1', 'Mar 2016', merge_format)

    arow = row+1
    worksheet.write(1, 0, "Company", title)
    worksheet.write(arow+2, 0, Company, title)

    for ele in range(6, 6*(len(shareHoldersFund)+1), 6):
        worksheet.merge_range(
            0, ele-5, 0, ele, updated_borrowings_dates_list[int((ele/6)-1)], merge_format)

        worksheet.set_column(ele-5, ele-5, 15)
        worksheet.write(1, ele-5, "Shareholder", title)
        worksheet.write(2, ele-5, "Funds", title)

        worksheet.write(1, ele-4, "Debts", title)

        worksheet.set_column(ele, ele, 10)
        worksheet.write(1, ele, "Cash", title)
        worksheet.write(2, ele, "Cash Eq", title)

        worksheet.write(
            arow+2, ele-5, shareHoldersFund[int((ele/6)-1)], title)
        worksheet.write(
            arow+2, ele-4, updated_borrowings_values_list[int((ele/6)-1)], title)

    for ele1 in range(6, 6*(len(totalRevenue)+1), 6):
        worksheet.merge_range(
            0, ele1-5, 0, ele1, updated_pldates_list[int((ele1/6)-1)], merge_format)

        worksheet.set_column(ele1-3, ele1-3, 15)
        worksheet.write(1, ele1-3, "Total", title)
        worksheet.write(2, ele1-3, "Revenue", title)

        worksheet.write(1, ele1-2, "PBT", title)

        worksheet.write(2, ele1-1, "PAT", title)

        worksheet.set_column(ele1, ele1, 10)
        worksheet.write(1, ele1, "Cash", title)
        worksheet.write(2, ele1, "Cash Eq", title)

        worksheet.write(arow+2, ele1-3, totalRevenue[int((ele1/6)-1)], title)
        worksheet.write(
            arow+2, ele1-2, updated_pbt_list[int((ele1/6)-1)], title)
        worksheet.write(
            arow+2, ele1-1, updated_pat_list[int((ele1/6)-1)], title)

    workbook.close()


if __name__ == '__main__':

    location = (r"C:\Users\jaypr\Desktop\Tech Stack\VSCodes\Web Scrapping\StockScrapping\Scrapping Screener Website Data\Stock Company List.xlsx")

    wb = openpyxl.load_workbook(location)
    sheet = wb.active
    link = []
    for i in range(1, 28):
        link.append(sheet.cell(row=i, column=1).value)

    for item in link:
        try:
            find_data(item, link.index(item))
        except:
            Continue
            print(f"An exception occurred with {item}")
