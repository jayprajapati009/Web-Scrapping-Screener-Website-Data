from ast import Continue
import time
from typing import final
from bs4 import BeautifulSoup as bs
from numpy import append
import requests
import xlsxwriter
import os
import openpyxl

# Setting up the directory to save the excel file in the same folder.
os.chdir(r'C:\Users\jaypr\Desktop\Tech Stack\VSCodes\Web Scrapping\StockScrapping\Scrapping Screener Website Data\Scraped Data')

location = (r'C:\Users\jaypr\Desktop\Tech Stack\VSCodes\Web Scrapping\StockScrapping\Scrapping Screener Website Data\Scraped Data\Merged_Data.xlsx')

workbook = xlsxwriter.Workbook('Merged_Data.xlsx')
worksheet = workbook.add_worksheet()
workbook.close()


def find_data(link, row):
    print(link)
    # To fetch the html data from the website
    html_text = requests.get(link).text

    # Parsing the data using lxml Parser and Beautiful Soup Library
    soup = bs(html_text, 'lxml')
    ##### Company Name #####
    Company = soup.find(
        'h1', class_='margin-0').text

    ##### Balance Sheet Months #####
    borrowings_dates_list = soup.find(
        'section', id='balance-sheet', class_='card card-large').table.thead.tr.text.strip().split('\n')

    ##### Borrowings or Debts - Balance Sheet #####
    borrowings_values = soup.find(
        'section', id='balance-sheet', class_='card card-large').text
    val_i = borrowings_values.find('Borrowings')
    othli_i = borrowings_values.find('Other Liabilities')
    borrowings_values_list = borrowings_values[val_i +
                                               10: othli_i].strip().split('\n')

    ##### Share Capital - Balance Sheet #####
    shareCap = soup.find(
        'section', id='balance-sheet', class_='card card-large').text
    shcap_i = shareCap.find('Share Capital')
    Res = shareCap.find('Reserves')
    shareCap_list = shareCap[shcap_i+15: Res].strip().split('\n')

    ##### Reserves - Balance Sheet #####
    Reserves = soup.find(
        'section', id='balance-sheet', class_='card card-large').text
    shcap_i = Reserves.find('Reserves')
    Res = Reserves.find('Borrowings')
    Reserves_list = Reserves[shcap_i+15: Res].strip().split('\n')

    ##### Profit and Loss Months #####
    pldates_list = soup.find(
        'section', id='balance-sheet', class_='card card-large').table.thead.tr.text.strip().split('\n')

    ##### Profit Before Tax (PBT) - Profit and Loss #####
    pbt = soup.find(
        'section', id='profit-loss', class_='card card-large').text
    pbt_i = pbt.find('Profit before tax')
    tax_i = pbt.find('Tax %')
    pbt_list = pbt[pbt_i+25: tax_i].strip().split('\n')

    ##### Net Profit (PAT) - Profit and Loss #####
    pat = soup.find(
        'section', id='profit-loss', class_='card card-large').text
    pat_i = pat.find('Net Profit')
    tax_i = pat.find('EPS')
    pat_list = pat[pat_i+15: tax_i].strip().split('\n')

    ##### Sales - Profit and Loss #####
    sales = soup.find(
        'section', id='profit-loss', class_='card card-large').text
    sales_i = sales.find('Sales\xa0+')
    tax_i = sales.find('Expenses\xa0+')
    sales_list = sales[sales_i+10: tax_i].strip().split('\n')

    ##### Other Income - Profit and Loss #####
    othInc = soup.find(
        'section', id='profit-loss', class_='card card-large').text
    othInc_i = othInc.find('Other')
    tax_i = othInc.find('Interest')
    othInc_list = othInc[othInc_i+15: tax_i].strip().split('\n')

    for slicingIndexb in range(len(borrowings_dates_list)):
        if "2016" in borrowings_dates_list[slicingIndexb]:
            break
        elif "2017" in borrowings_dates_list[slicingIndexb]:
            break
        elif "2018" in borrowings_dates_list[slicingIndexb]:
            break
        elif "2019" in borrowings_dates_list[slicingIndexb]:
            break
        elif "2020" in borrowings_dates_list[slicingIndexb]:
            break
        elif "2021" in borrowings_dates_list[slicingIndexb]:
            break
        elif "2022" in borrowings_dates_list[slicingIndexb]:
            break
        else:
            slicingIndexb = slicingIndexb + 1

    for slicingIndex in range(len(pldates_list)):
        if "2016" in pldates_list[slicingIndex]:
            break
        elif "2017" in pldates_list[slicingIndex]:
            break
        elif "2018" in pldates_list[slicingIndex]:
            break
        elif "2019" in pldates_list[slicingIndex]:
            break
        elif "2020" in pldates_list[slicingIndex]:
            break
        elif "2021" in pldates_list[slicingIndex]:
            break
        elif "2022" in pldates_list[slicingIndex]:
            break
        else:
            slicingIndex = slicingIndex + 1

    all_list = [borrowings_values_list,
                shareCap_list, Reserves_list, pbt_list, pat_list, sales_list, othInc_list]

    for ia in all_list:
        for ja in ia:
            ia[ia.index(ja)] = ja.replace(',', '')

    updated_borrowings_dates_list = borrowings_dates_list[slicingIndexb:len(
        borrowings_dates_list)]
    updated_borrowings_values_list = borrowings_values_list[slicingIndexb:len(
        borrowings_values)]
    updated_shareCap_list = shareCap_list[slicingIndexb:len(shareCap_list)]
    updated_Reserves_list = Reserves_list[slicingIndexb:len(Reserves_list)]
    shareHoldersFund = [float(updated_shareCap_list[i]) + float(updated_Reserves_list[i])
                        for i in range(len(updated_shareCap_list))]

    updated_pldates_list = pldates_list[slicingIndex:len(pldates_list)]
    updated_pbt_list = pbt_list[slicingIndex:len(pbt_list)]
    updated_pat_list = pat_list[slicingIndex:len(pat_list)]
    updated_sales_listt = sales_list[slicingIndex:len(sales_list)]
    updated_othInc_list = othInc_list[slicingIndex:len(othInc_list)]
    totalRevenue = [float(updated_sales_listt[i]) + float(updated_othInc_list[i])
                    for i in range(len(updated_sales_listt))]

    wb = openpyxl.load_workbook(location)
    sheet = wb.active

    arow = row+1
    sheet.cell(row=2, column=1).value = "Company"
    sheet.cell(row=arow+3, column=1).value = Company

    for ele in range(6, 6*(len(shareHoldersFund)+1), 6):
        sheet.merge_cells(start_row=1, start_column=ele -
                          4, end_row=1, end_column=ele+1)
        sheet.cell(
            1, ele-4).value = updated_borrowings_dates_list[int((ele/6)-1)]

        sheet.cell(2, ele-4).value = "Shareholder"
        sheet.cell(3, ele-4).value = "Funds"

        sheet.cell(2, ele-3).value = "Debts"

        sheet.cell(2, ele+1).value = "Cash"
        sheet.cell(3, ele+1).value = "Cash Eq"

        sheet.cell(
            arow+3, ele-4).value = shareHoldersFund[int((ele/6)-1)]
        sheet.cell(
            arow+3, ele-3).value = updated_borrowings_values_list[int((ele/6)-1)]

    if len(updated_pldates_list) != len(totalRevenue):
        for itt in range(abs(len(updated_pldates_list) - len(totalRevenue))):
            updated_pldates_list.append("TTM")

    for ele1 in range(6, 6*(len(totalRevenue)+1), 6):
        sheet.merge_cells(start_row=1, start_column=ele1 -
                          4, end_row=1, end_column=ele1+1)
        sheet.cell(
            1, ele1-4).value = updated_pldates_list[int((ele1/6)-1)]

        sheet.cell(2, ele1-2).value = "Total"
        sheet.cell(3, ele1-3).value = "Revenue"

        sheet.cell(2, ele1-1).value = "PBT"
        sheet.cell(2, ele1).value = "PAT"

        sheet.cell(
            arow+3, ele1-2).value = totalRevenue[int((ele1/6)-1)]
        sheet.cell(
            arow+3, ele1-1).value = updated_pbt_list[int((ele1/6)-1)]
        sheet.cell(
            arow+3, ele1).value = updated_pat_list[int((ele1/6)-1)]

    wb.save(location)


if __name__ == '__main__':

    location_2 = (
        r"C:\Users\jaypr\Desktop\Tech Stack\VSCodes\Web Scrapping\StockScrapping\Scrapping Screener Website Data\Stock Company List.xlsx")

    wb = openpyxl.load_workbook(location_2)
    sheet = wb.active
    link = []
    for i in range(1, 28):
        link.append(sheet.cell(row=i, column=1).value)

    for item in link:
        find_data(item, link.index(item))
        time_wait = 0.1
        time.sleep(time_wait)

        # try:
        #     find_data(item, link.index(item))
        #     time_wait = 0.1
        #     time.sleep(time_wait)
        # except:
        #     Continue
        #     print(f"An exception occurred with {item}")
